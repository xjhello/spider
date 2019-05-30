from openpyxl import load_workbook

# 打开一个workbook
wb = load_workbook("123.xlsx")

a=[]
b={}
# 获取当前活跃的worksheet,默认就是第一个worksheet
# ws = wb.active

# 当然也可以使用下面的方法

# 获取所有表格(worksheet)的名字
sheets = wb.sheetnames
# 第一个表格的名称
sheet_first = sheets[0]
# 获取特定的worksheet
ws = wb[sheet_first]

# 获取表格所有行和列，两者都是可迭代的
rows = ws.rows
columns = ws.columns

# 迭代所有的行
for row in rows:
    line = [col.value for col in row]
    key=str(line[0])+line[1]
    value = line[2]
    b = {key:value}
    a.append(b)
    # print(line)

# print(a)
fq = open('报考人数.txt', 'w', encoding='utf-8')
for line in a:
    # print(line)
    for index,key in enumerate(line):
        # print(key+' '+str(line[key]))
        fq.write(key+' '+str(line[key])+'\n')




